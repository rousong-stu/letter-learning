from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import UploadFile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession
import xlrd

from app.core.config import get_settings
from app.models import User
from app.models.word_book import WordBook, WordBookWord
from app.repositories.word_book import list_published_books as repo_list_books

settings = get_settings()

MAX_WORDS = 10000
SUPPORTED_HEADERS = {
    "word",
    "meaning_zh",
    "meaning_en",
    "example_en",
    "example_zh",
    "part_of_speech",
    "phonetic",
}


class WordBookUploadError(RuntimeError):
    pass


def _normalize_header(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_openpyxl_sheet(sheet: Worksheet) -> list[dict[str, str]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise WordBookUploadError("Excel 文件为空")
    headers_row = rows[0]
    header_map = {
        idx: _normalize_header(cell) for idx, cell in enumerate(headers_row)
    }
    if "word" not in header_map.values():
        raise WordBookUploadError("Excel 首行需包含 'word' 列")

    data: list[dict[str, str]] = []
    for row in rows[1:]:
        if row is None:
            continue
        item = {}
        for idx, cell in enumerate(row):
            header = header_map.get(idx)
            if header in SUPPORTED_HEADERS:
                item[header] = _normalize_cell(cell)
        word = item.get("word", "")
        if not word:
            continue
        data.append(item)
        if len(data) >= MAX_WORDS:
            break
    return data


def _parse_excel_bytes(content: bytes, filename: str | None) -> list[dict[str, str]]:
    suffix = (Path(filename or "").suffix or "").lower()
    if suffix not in {".xlsx", ".xls"}:
        raise WordBookUploadError("仅支持 .xlsx / .xls 格式")
    if suffix == ".xlsx":
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        return _parse_openpyxl_sheet(sheet)
    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        raise WordBookUploadError("Excel 文件为空")
    headers = {
        idx: _normalize_header(sheet.cell_value(0, idx))
        for idx in range(sheet.ncols)
    }
    if "word" not in headers.values():
        raise WordBookUploadError("Excel 首行需包含 'word' 列")
    data: list[dict[str, str]] = []
    for row_idx in range(1, sheet.nrows):
        item: dict[str, str] = {}
        for col_idx in range(sheet.ncols):
            header = headers.get(col_idx)
            if header in SUPPORTED_HEADERS:
                item[header] = _normalize_cell(sheet.cell_value(row_idx, col_idx))
        word = item.get("word", "")
        if not word:
            continue
        data.append(item)
        if len(data) >= MAX_WORDS:
            break
    return data


async def _save_cover_file(file: UploadFile | None) -> str | None:
    if not file:
        return None
    suffix = Path(file.filename or "").suffix.lower() or ".jpg"
    cover_dir = Path(settings.media_path) / "word_books" / "covers"
    cover_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{uuid4().hex}{suffix}"
    file_path = cover_dir / filename
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    return f"{settings.media_url}/word_books/covers/{filename}"


async def create_word_book(
    session: AsyncSession,
    *,
    title: str,
    description: str,
    language: str,
    level: str,
    tags_raw: str,
    cover_file: UploadFile | None,
    word_file: UploadFile,
    operator: User,
) -> WordBook:
    del operator  # 预留用于审计
    try:
        tags = json.loads(tags_raw) if tags_raw else []
        if not isinstance(tags, list):
            raise ValueError
    except ValueError as exc:
        raise WordBookUploadError("标签格式不正确，应为 JSON 数组") from exc

    word_file_bytes = await word_file.read()
    words_data = _parse_excel_bytes(word_file_bytes, word_file.filename)
    if not words_data:
        raise WordBookUploadError("解析到的单词为空，请检查文件内容")

    cover_url = await _save_cover_file(cover_file)

    book = WordBook(
        title=title.strip(),
        description=description.strip() or None,
        language=language or "en",
        level=level or None,
        tags=tags,
        cover_url=cover_url,
        total_words=len(words_data),
        is_published=True,
    )

    seen_words: set[str] = set()
    deduped_words: list[WordBookWord] = []
    skipped = 0
    for index, item in enumerate(words_data):
        word_value = item.get("word", "").strip()
        normalized = word_value.lower()
        if normalized in seen_words:
            skipped += 1
            continue
        seen_words.add(normalized)
        deduped_words.append(
            WordBookWord(
                word=word_value,
                meaning_zh=item.get("meaning_zh"),
                meaning_en=item.get("meaning_en"),
                example_en=item.get("example_en"),
                example_zh=item.get("example_zh"),
                part_of_speech=item.get("part_of_speech"),
                phonetic=item.get("phonetic"),
                order_index=len(deduped_words),
            )
        )

    book.total_words = len(deduped_words)
    book.words = deduped_words

    if skipped:
        book.description = (
            f"{book.description or ''}\n(已自动跳过 {skipped} 个重复单词)"
        ).strip()

    session.add(book)
    await session.flush()
    await session.refresh(book)
    return book


async def list_word_books(session: AsyncSession) -> list[WordBook]:
    return await repo_list_books(session)
